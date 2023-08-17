from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By

import pandas as pd

from openpyxl import Workbook, load_workbook
import requests

import time
import os

file_path = 'animeinfo.xlsx'
cnt = 6

def save_to_excel(sheet_name):
    book = Workbook()

    sheet = book.create_sheet(sheet_name, 0)
    sheet.append(['순번', '이름', '소개글', '테마', '사용자평 수', '평점'])
    book.save(file_path)

if not os.path.exists(file_path):
    save_to_excel('뽀로로')


def get_data():
    web = webdriver.Chrome()
    web.get('https://laftel.net/search?keyword=%EB%BD%80%EB%A1%9C%EB%A1%9C')
    time.sleep(1)
    elements = web.find_elements(By.TAG_NAME, 'img')
    real_cnt = 0
    for ele in elements:
        ele.click()
        time.sleep(1)
        btn = web.find_element(By.XPATH, '//*[@id="item-modal"]/div[1]/div/div[2]/div/div/button')
        btn.click()
        review = web.find_element(By.XPATH,'//*[@id="item-modal"]/div[2]/div[1]/div[1]/a[2]')
        review.click()
        time.sleep(1)
        parse_data(web.page_source,int(real_cnt)+1)
        time.sleep(1)
        web.find_element(By.XPATH, '/html/body/div[1]/div/div[3]/div[2]/div/div[1]/div/div[2]/nav/button[2]').click()
        time.sleep(1)
        real_cnt +=1
        if real_cnt == cnt:
            break


    # 제목 줄거리 더보기 평균별점 사용자평개수
    # 옵션 이미지수
    web.quit()

    # with open('index.html', "w",encoding='utf-8') as file:
    #     file.write(web.page_source)


def parse_data(html,real_cnt):
    bs = BeautifulSoup(html, 'html.parser')
    h1 = bs.find('h1')
    summary = bs.find('summary')
    ach = bs.select('ul li a')
    auser = bs.select('div > div a')
    avg = bs.select('div>section>div>div>div')

    aarr = []
    usercnt =0
    for a in ach:
        if '#' in a.getText():
            aarr.append(a.getText().split('#')[1])

    # print(auser)
    for a in auser:
        span = a.find('span')
        if span is not None:
            if not '더빙' in span.getText():
                # print(span)
                usercnt = span.getText()
    for div in avg:
        if '.' in div.getText():
            # print(span)
            avg = div.getText()
            break

    print('제목',h1.getText())
    print('줄거리',summary.getText())
    print('테마',aarr)
    print('사용자평개수',usercnt)
    print('평점', avg)

    append_to_excel([real_cnt,h1.getText(),summary.getText(),' '.join(aarr),usercnt,avg])


    # ['순번', '이름', '소개글', '테마', '사용자평 수', '평점']
    # for li in lis:
    #     a = li.find('a')
    #     print(a)
    #     title = a.find('p').get_text()


def append_to_excel(data):
    book = load_workbook('animeinfo.xlsx')

    first_sheet = book.worksheets[0]

    first_sheet.append(data)

    book.save('animeinfo.xlsx')


if __name__ == '__main__':
    get_data()
